import sqlite3 as sq
import openpyxl
from openpyxl.styles import PatternFill
import time

time_start = time.time()
print("Час початку: ")
print(time.ctime(time_start)[10:19])


# Створення нової таблиці, це не обов'язково.

# con = sq.connect('customers.db')
# cur = con.cursor()
# cur.execute("CREATE TABLE IF NOT EXISTS customers(name TEXT, egrpou INTEGER, dealer_code TEXT)") #custoners table
# cur.execute("CREATE TABLE IF NOT EXISTS dealers(name TEXT, egrpou INTEGER, code TEXT, percentages INTEGER)")#dealers table
# con.close()


# ФУНКЦІЇ ДЛЯ ВИВАНТАЖЕННЯ КЛІЄНТІВ ДИЛЕРІВ

# Вивантаження усіх даних про клієнтів дилерів. Створює файл "customer_data.xlsx"
def export_all_customers_data():
    # excel
    book = openpyxl.Workbook()
    sheet = book.active
    con = sq.connect('customers.db')
    cur = con.cursor()
    cur.execute(
        "SELECT customers.name, customers.egrpou, dealers.code FROM customers JOIN dealers ON dealer_code == code")
    test2 = cur.fetchall()
    count = 0
    for i in test2:
        count += 1
        sheet["A" + str(count)] = i[0]
        sheet["B" + str(count)] = i[1]
        sheet["C" + str(count)] = i[2]
        print(i[0])
    print(test2)
    book.save("customer_data.xlsx")
    # closing DB
    con.close()


# export_all_customers_data()

# Вивантаження даних про клієнтів певного дилера. Створює файл "customer_data.xlsx"
def export_customers_data_by_code(code: str):  # code - код дилера
    # excel open
    book = openpyxl.Workbook()
    sheet = book.active
    code = code.strip()
    print(code)
    con = sq.connect('customers.db')
    cur = con.cursor()
    cur.execute(
        "SELECT customers.name, customers.egrpou, customers.dealer_code FROM customers WHERE customers.dealer_code == :code",
        {'code': code})
    test2 = cur.fetchall()

    count = 0
    print(len(test2))
    for i in range(len(test2)):
        count += 1
        sheet["A" + str(count)] = test2[i][0]
        sheet["B" + str(count)] = test2[i][1]
        sheet["C" + str(count)] = test2[i][2]

    print(test2)
    book.save("customer_data.xlsx")

    # closing DB
    con.close()


# export_customers_data_by_code("КОД ПАРТНЕРА")


# ОКРЕМІ ФУНКЦІЇ ДЛЯ ВИКОНАННЯ ГРУПУВАННЯ ДАНИХ ПО ОПЛАТАМ КЛІЄНТІВ ДИЛЕРІВ


# відкриття файлу з даними про оплати клієнтів (вигрузка з Терасофт), порівняння з базою клієнтів дилерів; знайдені оплати зберігаються у вигляді словника
def paid_customer_count():
    # path = str(input("Введи шлях до вигрузки з оплатами клієнтів (файл у форматі .xlsx: "))
    path = r'C:\Users\Vadym\Documents\projects\dealers_customers\usrrahunok_03_10_2022_15_55.xlsx' #шлях до файлу з даними оплат
    book_cash = openpyxl.open(path)
    sheet_cash = book_cash.active
    con = sq.connect('customers.db')
    cur = con.cursor()
    cur.execute("SELECT * FROM customers")
    customers_list = cur.fetchall()
    result_data = []
    temp_arr = []
    count_pay = 0
    count_without_code = 0
    count_for_array = 0

    try:
        for i in range(2403, sheet_cash.max_row + 1): #sheet_cash.max_row + 1
            print(f"перевірка рядка: {i}")
            if sheet_cash[i][9].value == None and len(str(sheet_cash[i][0].value)) > 0:
                print("Оплата без коду ЄДРПОУ. Пошукай такі оплати по фільтрам в Ексель та перевір руками.")
                count_without_code += 1
                continue
            elif len(str(sheet_cash[i][0].value)) < 0:
                print("Кінець файлу")
                break
            for j in customers_list:
                if int(sheet_cash[i][9].value) == int(j[1]):
                    print("Знайшов оплату!")
                    count_pay += 1
                    temp_arr.append(j[0]) #ПІБ
                    temp_arr.append(sheet_cash[i][7].value) #сума
                    temp_arr.append(sheet_cash[i][0].value) #дата оплати
                    temp_arr.append(j[2]) #код партнера
                    temp_arr.append(j[1]) #єдрпоу клієнта

                    result_data.append(temp_arr)
                    temp_arr = []
                    print(result_data)
                    print(f"Це оплата (з кодом ЄДРПОУ): № {count_pay}")
    except Exception as ex:
        print(f"вилізла помилка в функції paid_customer_count, на рядку №{i}")
        print(ex)
        # print(result_data)
    finally:
        book_cash.close()
        con.close()
        print(f"ЗАГАЛЬНА КІЛЬКІСТЬ ОПЛАТ З КОДОМ ЄДРПОУ: {count_pay}")
        print(f"ЗАГАЛЬНА КІЛЬКІСТЬ ОПЛАТ З БЕЗ КОДУ ЄДРПОУ: {count_without_code}")
        return result_data


# print(paid_customer_count())
result_paid_customer_count = paid_customer_count()


# розширення даних про дилерів, а саме - додає єдрпоу, назву та ін. до результатів функції 'paid_customer_count'; для цього треба виконання попередньо функції 'paid_customer_count'
def add_more_diler_info(data):  # вказати результат з функції 'paid_customer_count'
    paid_customers_arr = data
    con = sq.connect('customers.db')
    cur = con.cursor()
    cur.execute("SELECT * FROM dealers")
    dealers_data = cur.fetchall()
    try:
        for arr in range(len(paid_customers_arr)):
            print(paid_customers_arr[arr])

            for code in dealers_data:
                if paid_customers_arr[arr][3].lower() in code[2].lower():
                    paid_customers_arr[arr].append(code[0])
                    paid_customers_arr[arr].append(code[1])
                    paid_customers_arr[arr].append(code[3])
            print("повні дані: ",paid_customers_arr[arr])
    except:
        print(f"вилізла помилка у функції add_more_diler_info, в масиві: {arr}")
    finally:
        con.close()
        print(paid_customers_arr)
        return paid_customers_arr


result_add_more_diler_info = add_more_diler_info(result_paid_customer_count)

# вивантаження даних роботи функції 'add_more_diler_info' у файл Ексель; дані будуть збережені у файл 'result_paid_customers.xlsx'; дані, які треба продублювати для Кешдеска - підкреслюються
def paste_info_to_result_excel(data: list):  # вставити в data результат функції 'add_more_diler_info'

    book_result = openpyxl.open('result_paid_customers.xlsx') #ексель-файл шаблон, у котрому будуть зберігатися результати
    sheet_result = book_result.active
    count = 2
    my_fill = PatternFill(start_color="FFFFFA00", fill_type="solid")
    try:
        for arr in range(len(data)):
            sheet_result[count][0].value = data[arr][1]  # suma
            sheet_result[count][1].value = data[arr][0]  # customer name
            sheet_result[count][2].value = data[arr][5]  # dealer name
            sheet_result[count][3].value = data[arr][6]  # dealer edrpou
            if data[arr][6] == 3345416704:
                for i in range(8):
                    sheet_result[count][i].fill = my_fill
            sheet_result[count][4].value = data[arr][2]  # pay date
            sheet_result[count][5].value = data[arr][4]  # customer edrpou
            sheet_result[count][6].value = data[arr][7]  # percentage
            sheet_result[count][7].value = "E-receipt"
            count += 1
    except Exception as ex:
        print(f"вилізла помилка в функції paste_info_to_result_excel, на рядку: {count}")
        print(ex)

    finally:
        book_result.save('result_paid_customers.xlsx')
        book_result.close()

        time_end = time.time() - time_start
        print(f"Час закінчення: {time.ctime(time.time())[10:19]}")
        print(f"Програма працювала: {time.gmtime(time_end)[3]} годин {time.gmtime(time_end)[4]} хвилин {time.gmtime(time_end)[5]} секунд")


paste_info_to_result_excel(result_add_more_diler_info)

